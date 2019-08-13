import tkinter
from tkinter import filedialog
import mysql.connector 
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, colors
from openpyxl.styles.borders import Border, Side
# Window Configure
wd = tkinter.Tk()
wd.title("Schema-Exporter")
wd.resizable(False,False)
wd.geometry("300x500")


tborder = Border(top=Side(style='thin'),bottom=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'))
_dbList = []
_SelectdbList = []
# Function
def exportSchema():
    _svrhost = en1.get()
    _svrport = en2.get()
    _svruser = en3.get()
    _svrpass = en4.get()
    try:
        _dbConfig = {'user':_svruser,'password':_svrpass,'host':_svrhost,'port':_svrport,'connection_timeout':10}
        _dbCon = mysql.connector.connect(**_dbConfig)
        _dbCur = _dbCon.cursor(dictionary=True)
        _dbCur.execute("SELECT table_schema FROM information_schema.TABLES WHERE table_schema NOT IN ('information_schema','mysql','performance_schema') GROUP BY table_schema")
        _dbData = _dbCur.fetchall()        
        _dbCur.close()
        _dbCon.close()
       
        # Disabled Box
        sb1.config(state="disabled")
        lb5.config(text="Connection Ok.")
        en1.config(state="disabled")
        en2.config(state="disabled")
        en3.config(state="disabled")
        en4.config(state="disabled")
        sb2.config(state="normal")
        for _schema in _dbData:
            _dbList.append(_schema['table_schema'])
        listvar = tkinter.Variable()
        listvar.set(tuple(_dbList))
        listbox.config(listvariable=listvar)
    except Exception as DBERR:
        print(DBERR)
        lb5.config(text="Connection Fail.")

def saveSchema():   
    # excel Configure
    exwb = Workbook()

    # Target Schema Index
    for _listData in listbox.curselection():
        _SelectdbList.append(_dbList[_listData])

    # Sheet Create 
    for _scnm in _SelectdbList:
        globals()[_scnm] = exwb.create_sheet(_scnm)
    
    # File Name and Path
    filepath =  filedialog.askdirectory (initialdir ="C:/",title = "choose your file")
    filename = str(filepath) + "/" + "Schema-Exporter_" + str(en1.get()) + ".xlsx"

    # Connection Configure
    _svrhost = en1.get()
    _svrport = en2.get()
    _svruser = en3.get()
    _svrpass = en4.get()
    try:
        _dbConfig = {'user':_svruser,'password':_svrpass,'host':_svrhost,'port':_svrport,'connection_timeout':10}
        _dbCon = mysql.connector.connect(**_dbConfig)
        _dbCur = _dbCon.cursor(dictionary=True)
        for _chschema in _SelectdbList:
            _Query = "SELECT table_schema,table_name,table_type,ENGINE,row_format,table_collation,table_comment FROM information_schema.TABLES WHERE table_schema='%s'" % _chschema
            _dbCur.execute(_Query)
            _dbData = _dbCur.fetchall()
            
            # Reset Row
            rowstrpnt = 1
            str_row = 0

            for _tbnm in _dbData:                
                _dtlQuery = "SELECT column_name,column_default,is_nullable,data_type,character_maximum_length AS column_length,character_set_name,collation_name,column_key,extra,column_comment FROM information_schema.COLUMNS WHERE table_name='%s' AND table_schema='%s' ORDER BY ordinal_position;" % (_tbnm['table_name'],_tbnm['table_schema'])
                _dbCur.execute(_dtlQuery)
                _dbDatadtl = _dbCur.fetchall()
                str_row = rowstrpnt
                # Table Name
                globals()[_chschema].merge_cells(start_row=rowstrpnt,start_column=1,end_row=rowstrpnt,end_column=2)
                globals()[_chschema].merge_cells(start_row=rowstrpnt,start_column=3,end_row=rowstrpnt,end_column=10)
         
                globals()[_chschema]['A' + str(rowstrpnt)] = "Table Name"
                globals()[_chschema]['C' + str(rowstrpnt)] = _tbnm["table_name"]
                # Bold
                globals()[_chschema]['A' + str(rowstrpnt)].font = Font(bold=True)
                globals()[_chschema]['A' + str(rowstrpnt)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                # Table Description
                globals()[_chschema].merge_cells(start_row=rowstrpnt+1,start_column=1,end_row=rowstrpnt+1,end_column=2)
                globals()[_chschema].merge_cells(start_row=rowstrpnt+1,start_column=3,end_row=rowstrpnt+1,end_column=10)
                globals()[_chschema]['A' + str(rowstrpnt + 1)] = "Description"
                globals()[_chschema]['C' + str(rowstrpnt + 1)] = _tbnm["table_comment"]

                # Bold
                globals()[_chschema]['A' + str(rowstrpnt + 1)].font = Font(bold=True)
                globals()[_chschema]['A' + str(rowstrpnt + 1)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                # Column Feild
                globals()[_chschema]['A' + str(rowstrpnt + 2)] = "No"
                globals()[_chschema]['B' + str(rowstrpnt + 2)] = "Column Name"
                globals()[_chschema]['C' + str(rowstrpnt + 2)] = "Data Type"
                globals()[_chschema]['D' + str(rowstrpnt + 2)] = "Length"
                globals()[_chschema]['E' + str(rowstrpnt + 2)] = "Nullable"
                globals()[_chschema]['F' + str(rowstrpnt + 2)] = "Key"
                globals()[_chschema]['G' + str(rowstrpnt + 2)] = "Extra"
                globals()[_chschema]['H' + str(rowstrpnt + 2)] = "Collation"
                globals()[_chschema]['I' + str(rowstrpnt + 2)] = "Default"     
                globals()[_chschema]['J' + str(rowstrpnt + 2)] = "Comment"         
                
                # Bold
                globals()[_chschema]['A' + str(rowstrpnt + 2)].font = Font(bold=True)
                globals()[_chschema]['B' + str(rowstrpnt + 2)].font = Font(bold=True)
                globals()[_chschema]['C' + str(rowstrpnt + 2)].font = Font(bold=True)
                globals()[_chschema]['D' + str(rowstrpnt + 2)].font = Font(bold=True)
                globals()[_chschema]['E' + str(rowstrpnt + 2)].font = Font(bold=True)
                globals()[_chschema]['F' + str(rowstrpnt + 2)].font = Font(bold=True)
                globals()[_chschema]['G' + str(rowstrpnt + 2)].font = Font(bold=True)
                globals()[_chschema]['H' + str(rowstrpnt + 2)].font = Font(bold=True)
                globals()[_chschema]['I' + str(rowstrpnt + 2)].font = Font(bold=True)
                globals()[_chschema]['J' + str(rowstrpnt + 2)].font = Font(bold=True)

                globals()[_chschema]['A' + str(rowstrpnt + 2)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                globals()[_chschema]['B' + str(rowstrpnt + 2)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                globals()[_chschema]['C' + str(rowstrpnt + 2)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                globals()[_chschema]['D' + str(rowstrpnt + 2)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                globals()[_chschema]['E' + str(rowstrpnt + 2)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                globals()[_chschema]['F' + str(rowstrpnt + 2)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                globals()[_chschema]['G' + str(rowstrpnt + 2)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                globals()[_chschema]['H' + str(rowstrpnt + 2)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                globals()[_chschema]['I' + str(rowstrpnt + 2)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid") 
                globals()[_chschema]['J' + str(rowstrpnt + 2)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                # Row Set
                rowstrpnt += 3
                for idx,_rows in enumerate(_dbDatadtl):
                    globals()[_chschema]['A' + str(rowstrpnt)] = idx +1
                    globals()[_chschema]['B' + str(rowstrpnt)] = _rows['column_name']
                    globals()[_chschema]['C' + str(rowstrpnt)] = _rows['data_type']
                    globals()[_chschema]['D' + str(rowstrpnt)] = _rows['column_length']
                    globals()[_chschema]['E' + str(rowstrpnt)] = _rows['is_nullable']
                    globals()[_chschema]['F' + str(rowstrpnt)] = _rows['column_key']
                    globals()[_chschema]['G' + str(rowstrpnt)] = _rows['extra']
                    globals()[_chschema]['H' + str(rowstrpnt)] = _rows['collation_name']
                    globals()[_chschema]['I' + str(rowstrpnt)] = _rows['column_default']
                    globals()[_chschema]['J' + str(rowstrpnt)] = _rows['column_comment']
                    rowstrpnt += 1
                # Constraint Info
                globals()[_chschema].merge_cells(start_row=rowstrpnt,start_column=1,end_row=rowstrpnt,end_column=10)
                globals()[_chschema].merge_cells(start_row=rowstrpnt+1,start_column=1,end_row=rowstrpnt+1,end_column=3)
                globals()[_chschema].merge_cells(start_row=rowstrpnt+1,start_column=4,end_row=rowstrpnt+1,end_column=5)
                globals()[_chschema].merge_cells(start_row=rowstrpnt+1,start_column=6,end_row=rowstrpnt+1,end_column=10)
                globals()[_chschema]['A' + str(rowstrpnt)] = "Constraint Info"
                globals()[_chschema]['A' + str(rowstrpnt+1)] = "Name"
                globals()[_chschema]['D' + str(rowstrpnt+1)] = "Column"
                globals()[_chschema]['F' + str(rowstrpnt+1)] = "Referance"

                globals()[_chschema]['A' + str(rowstrpnt)].font = Font(bold=True)
                globals()[_chschema]['A' + str(rowstrpnt+1)].font = Font(bold=True)
                globals()[_chschema]['D' + str(rowstrpnt+1)].font = Font(bold=True)
                globals()[_chschema]['F' + str(rowstrpnt+1)].font = Font(bold=True)

                # globals()[_chschema]['A' + str(rowstrpnt)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                # globals()[_chschema]['A' + str(rowstrpnt+1)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                # globals()[_chschema]['D' + str(rowstrpnt+1)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                # globals()[_chschema]['F' + str(rowstrpnt+1)].fill = PatternFill(start_color='00d3d3d3',end_color='00d3d3d3', fill_type = "solid")
                rowstrpnt += 3
                

                _Conquery = """
                SELECT
                    constraint_name as constraint_key,
                    group_concat(column_name) as con_column,
                    concat(referenced_table_schema,'.',referenced_table_name,' : ',referenced_column_name) as refer_info
                FROM
                    information_schema.KEY_COLUMN_USAGE 
                WHERE
                    CONSTRAINT_SCHEMA = '%s' 
                    AND table_name = '%s'
                    AND constraint_name <> 'PRIMARY'
                GROUP BY
                    constraint_name
                """ % (_chschema,_tbnm["table_name"])
                _dbCur.execute(_Conquery)
                _dbDataCon = _dbCur.fetchall() 
                for _constraint in _dbDataCon:
                    
                    globals()[_chschema].merge_cells(start_row=rowstrpnt-1,start_column=1,end_row=rowstrpnt-1,end_column=3)
                    globals()[_chschema].merge_cells(start_row=rowstrpnt-1,start_column=4,end_row=rowstrpnt-1,end_column=5)
                    globals()[_chschema].merge_cells(start_row=rowstrpnt-1,start_column=6,end_row=rowstrpnt-1,end_column=10)
                    globals()[_chschema]['A' + str(rowstrpnt-1)] = _constraint['constraint_key']
                    globals()[_chschema]['D' + str(rowstrpnt-1)] = _constraint['con_column']
                    globals()[_chschema]['F' + str(rowstrpnt-1)] = _constraint['refer_info']
                    
                    rowstrpnt += 1
                # Table INFO
                globals()[_chschema].merge_cells(start_row=rowstrpnt,start_column=1,end_row=rowstrpnt,end_column=2)
                globals()[_chschema].merge_cells(start_row=rowstrpnt,start_column=3,end_row=rowstrpnt,end_column=4)
                globals()[_chschema]['A' + str(rowstrpnt)] = "Engine"
                globals()[_chschema]['C' + str(rowstrpnt)] = _tbnm["ENGINE"]
                globals()[_chschema]['A' + str(rowstrpnt)].font = Font(bold=True)
                globals()[_chschema].merge_cells(start_row=rowstrpnt,start_column=5,end_row=rowstrpnt,end_column=6)
                globals()[_chschema].merge_cells(start_row=rowstrpnt,start_column=7,end_row=rowstrpnt,end_column=10)
                globals()[_chschema]['E' + str(rowstrpnt)] = "Row Format"
                globals()[_chschema]['G' + str(rowstrpnt)] = _tbnm["row_format"]
                globals()[_chschema]['E' + str(rowstrpnt)].font = Font(bold=True)

                globals()[_chschema].merge_cells(start_row=rowstrpnt+1,start_column=1,end_row=rowstrpnt+1,end_column=2)
                globals()[_chschema].merge_cells(start_row=rowstrpnt+1,start_column=3,end_row=rowstrpnt+1,end_column=4)
                globals()[_chschema]['A' + str(rowstrpnt+1)] = "Table Type"
                globals()[_chschema]['C' + str(rowstrpnt+1)] = _tbnm["table_type"]
                globals()[_chschema]['A' + str(rowstrpnt+1)].font = Font(bold=True)
                globals()[_chschema].merge_cells(start_row=rowstrpnt+1,start_column=5,end_row=rowstrpnt+1,end_column=6)
                globals()[_chschema].merge_cells(start_row=rowstrpnt+1,start_column=7,end_row=rowstrpnt+1,end_column=10)
                globals()[_chschema]['E' + str(rowstrpnt+1)] = "Collation"
                globals()[_chschema]['G' + str(rowstrpnt+1)] = _tbnm["table_collation"]
                globals()[_chschema]['E' + str(rowstrpnt+1)].font = Font(bold=True)

                for col in range(1,11):
                    for row in range(str_row,rowstrpnt+2):
                        globals()[_chschema].cell(row=row,column=col).border = tborder


                rowstrpnt += 6

        _dbCur.close()
        _dbCon.close()
        
        lb7.config(text="Export Success.")
        sb2.config(state="disabled")


        # Workbook Save
        exwb.remove(exwb['Sheet'])
        exwb.save(filename)

    except Exception as DBERROR:
        lb7.config(text="Export Fail.")




# Label
lb1 = tkinter.Label(wd,text="Connect Server : ")
lb2 = tkinter.Label(wd,text="Port : ")
lb3 = tkinter.Label(wd,text="Account : ")
lb4 = tkinter.Label(wd,text="Password : ")
lb5 = tkinter.Label(wd,text="")
lb6 = tkinter.Label(wd,text="Choose Database : ")
lb7 = tkinter.Label(wd,text="")
lb8 = tkinter.Label(wd,text="Create by ISEDBA Park Sang Jea")
# Input 
en1 = tkinter.Entry(wd)
en2 = tkinter.Entry(wd)
en3 = tkinter.Entry(wd)
en4 = tkinter.Entry(wd)

# Button
sb1 = tkinter.Button(wd,text="Connect Server", overrelief="solid", width=15, command=exportSchema, repeatdelay=1000, repeatinterval=100,state="normal")
sb2 = tkinter.Button(wd,text="Export Schema", overrelief="solid", width=15, command=saveSchema, repeatdelay=1000, repeatinterval=100,state="disabled")


# List
listbox = tkinter.Listbox(wd, selectmode='extended', height=0)

# Place
## Server
lb1.place(x=10,y=10)
en1.place(x=150,y=10)
## Port
lb2.place(x=10,y=30)
en2.place(x=150,y=30)
## Acc
lb3.place(x=10,y=50)
en3.place(x=150,y=50)
## Pass
lb4.place(x=10,y=70)
en4.place(x=150,y=70)


sb1.place(x=10,y=100)
lb5.place(x=150,y=100)
lb6.place(x=10,y=130)
sb2.place(x=10,y=190)
lb7.place(x=10,y=220)
listbox.place(x=150,y=130)



lb8.place(x=10,y=480)





# Run UI
wd.mainloop()


